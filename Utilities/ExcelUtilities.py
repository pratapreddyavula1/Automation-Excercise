import openpyxl
from openpyxl.styles import PatternFill

"""Reading Excel Data for data driven testing"""


class ExcelUtils:
    @staticmethod
    def get_rows(file):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        return sheet.max_row

    @staticmethod
    def get_columns(file):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        return sheet.max_column

    @staticmethod
    def read_data(file, rownum, colnum):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        return sheet.cell(row=rownum, column=colnum).value

    @staticmethod
    def write_data(file, rownum, colnum, data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        sheet.cell(row=rownum, column=colnum).value = data
        workbook.save(file)

    @staticmethod
    def green_color(file, rownum, colnum):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        fill_green = PatternFill(start_color='60b212', end_color='60b212', fill_type='solid')
        sheet.cell(row=rownum, column=colnum).fill = fill_green
        workbook.save(file)

    @staticmethod
    def red_color(file, rownum, colnum):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        fill_red = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
        sheet.cell(row=rownum, column=colnum).fill = fill_red
        workbook.save(file)
